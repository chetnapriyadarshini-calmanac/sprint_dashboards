"""
capacity_excel.py
-----------------
Read per-member sprint capacity from the maintained capacity workbook.

`CAPACITY_XLSX` (in sprint_dashboard_config.py) may point at any of:

  * a LOCAL .xlsx path (e.g. the committed Team_Capacity.xlsx, or a Drive/
    OneDrive synced file), or
  * a GOOGLE SHEET URL, read via one of (tried in this order):
        1. a service-account key   (CAPACITY_SA_KEY / GOOGLE_APPLICATION_CREDENTIALS)
        2. YOUR Google account via OAuth  (CAPACITY_OAUTH_CLIENT + cached token)
        3. an unauthenticated public download (only if the link is public)

Capacity model:

    Sprint cap(row) = Capacity/day x (Working days - Team days off - Days off)

We compute this from the raw inputs (Capacity/day, Days off + the Settings
Working days / Team days off) rather than trusting the workbook's cached formula
value, so it is correct even right after a hand-edit that hasn't recalculated
yet. Multiple activity rows per member are summed downstream by
compute_capacity() in the generator.

Returns the canonical shape the generator consumes:
    Member | Activity | Sprint cap

CLI:
    python capacity_excel.py            # prints the resolved capacity table
"""

from __future__ import annotations

import io
import os
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

DEFAULT_FILE = "Team_Capacity.xlsx"

# Read-only Drive scope (enough to export a Sheet to .xlsx).
_SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]

# Default credential file names, tried when the config values are None.
# All are git-ignored (see .gitignore).
_DEFAULT_SA_KEY_NAMES = (".gcp_sa.json", "service_account.json")
_DEFAULT_OAUTH_CLIENT_NAMES = (".gcp_oauth_client.json", "client_secret.json")
_DEFAULT_TOKEN_NAME = ".gcp_oauth_token.json"

# .xlsx files are ZIP archives - they start with the "PK" local-file header.
_XLSX_MAGIC = b"PK\x03\x04"
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _num(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _is_url(src: str) -> bool:
    return src.lower().startswith(("http://", "https://"))


def _extract_sheet_id(url: str) -> str | None:
    """Return the Google Sheet file ID from a share/edit/export URL, else None."""
    if "docs.google.com/spreadsheets" not in url:
        return None
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
    return m.group(1) if m else None


def _normalise_gsheet_url(url: str) -> str:
    """Turn a Google Sheets share/edit link into an .xlsx export URL (used only
    for the UNAUTHENTICATED public-link path). Non-Sheet URLs pass through."""
    sid = _extract_sheet_id(url)
    if not sid or "export?format=" in url or "/export?" in url:
        return url
    return f"https://docs.google.com/spreadsheets/d/{sid}/export?format=xlsx"


def _first_existing(candidates) -> str | None:
    for c in candidates:
        if c and Path(c).exists():
            return str(c)
    return None


def _find_sa_key(explicit: str | None) -> str | None:
    """service-account key: explicit -> GOOGLE_APPLICATION_CREDENTIALS -> default
    file at the repo root or next to this script. None if not found."""
    here = Path(__file__).resolve().parent
    cands: list = [explicit, os.environ.get("GOOGLE_APPLICATION_CREDENTIALS")]
    for name in _DEFAULT_SA_KEY_NAMES:
        cands += [here.parent / name, here / name]
    return _first_existing(cands)


def _find_oauth_client(explicit: str | None) -> str | None:
    """OAuth client-secret file: explicit -> default name at the repo root or
    next to this script. None if not found."""
    here = Path(__file__).resolve().parent
    cands: list = [explicit]
    for name in _DEFAULT_OAUTH_CLIENT_NAMES:
        cands += [here.parent / name, here / name]
    return _first_existing(cands)


def _default_token_path(client_path: str, explicit: str | None) -> str:
    """Where to cache the OAuth token — explicit if given, else next to the
    client-secret file."""
    if explicit:
        return explicit
    return str(Path(client_path).resolve().parent / _DEFAULT_TOKEN_NAME)


def _drive_export_xlsx(creds, file_id: str) -> io.BytesIO:
    """Fetch a Drive file as .xlsx bytes with an authorized credential.

    Handles BOTH kinds of file that a docs.google.com/spreadsheets URL can point
    at: a NATIVE Google Sheet (exported to .xlsx) and an UPLOADED .xlsx file
    (downloaded as-is — the Drive export API rejects those with
    'Export only supports Docs Editors files')."""
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload

    svc = build("drive", "v3", credentials=creds, cache_discovery=False)
    meta = svc.files().get(fileId=file_id, fields="mimeType,name",
                           supportsAllDrives=True).execute()
    mime = meta.get("mimeType", "")

    if mime == "application/vnd.google-apps.spreadsheet":
        # Native Google Sheet -> export to .xlsx
        data = svc.files().export(fileId=file_id, mimeType=_XLSX_MIME).execute()
        buf = io.BytesIO(data if isinstance(data, (bytes, bytearray))
                         else str(data).encode("utf-8", "ignore"))
    else:
        # Uploaded binary file (.xlsx etc.) -> download the bytes directly
        buf = io.BytesIO()
        dl = MediaIoBaseDownload(
            buf, svc.files().get_media(fileId=file_id, supportsAllDrives=True))
        done = False
        while not done:
            _, done = dl.next_chunk()

    raw = buf.getvalue()
    if not raw or not raw.startswith(_XLSX_MAGIC):
        raise RuntimeError(
            f"Fetched Drive file (mimeType={mime!r}) is not a valid .xlsx "
            "workbook. Point CAPACITY_XLSX at a Google Sheet or an .xlsx file you "
            "can open, with 'Settings' and 'Capacity' tabs."
        )
    buf.seek(0)
    return buf


def _export_gsheet_via_service_account(sheet_id: str, key_path: str) -> io.BytesIO:
    """Export a Sheet using a service account (Sheet shared with its client_email)."""
    try:
        from google.oauth2 import service_account
    except ImportError as e:
        raise RuntimeError(
            "Missing Google libraries. Run:\n"
            "  pip install google-api-python-client google-auth\n"
            f"(import error: {e})"
        )
    try:
        creds = service_account.Credentials.from_service_account_file(key_path, scopes=_SCOPES)
        return _drive_export_xlsx(creds, sheet_id)
    except Exception as e:
        raise RuntimeError(
            f"Could not export the Google Sheet (id={sheet_id}) with the service "
            f"account key '{key_path}'.\n"
            "Check: (1) Sheet shared with the service account client_email (Viewer); "
            "(2) Drive API enabled on its project; (3) key file valid.\n"
            f"Underlying error: {e}"
        )


def _export_gsheet_via_oauth(sheet_id: str, client_path: str, token_path: str) -> io.BytesIO:
    """Export a Sheet as the SIGNED-IN USER via OAuth.

    First run opens a browser for consent (sign in with the Google account that
    can see the Sheet); the resulting token is cached at `token_path` and
    refreshed automatically on later runs, so subsequent runs are headless.
    """
    try:
        from google.oauth2.credentials import Credentials
        from google_auth_oauthlib.flow import InstalledAppFlow
        from google.auth.transport.requests import Request
    except ImportError as e:
        raise RuntimeError(
            "Missing Google OAuth libraries. Run:\n"
            "  pip install google-api-python-client google-auth google-auth-oauthlib\n"
            f"(import error: {e})"
        )

    creds = None
    if Path(token_path).exists():
        try:
            creds = Credentials.from_authorized_user_file(token_path, _SCOPES)
        except Exception:
            creds = None

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
            except Exception:
                creds = None
        if not creds or not creds.valid:
            # Interactive consent (browser). Needed once, then the cached token
            # is reused/refreshed.
            flow = InstalledAppFlow.from_client_secrets_file(client_path, _SCOPES)
            creds = flow.run_local_server(port=0)
        try:
            Path(token_path).write_text(creds.to_json(), encoding="utf-8")
        except Exception:
            pass  # non-fatal: we simply re-consent next time

    try:
        return _drive_export_xlsx(creds, sheet_id)
    except Exception as e:
        raise RuntimeError(
            f"Could not export the Google Sheet (id={sheet_id}) as the signed-in "
            f"user.\nCheck that your account can open the Sheet and the Drive API "
            f"is enabled on the OAuth client's project.\nUnderlying error: {e}"
        )


def _fetch_workbook_bytes(url: str) -> io.BytesIO:
    """UNAUTHENTICATED download from a URL. Only works for public links."""
    import requests

    fetch_url = _normalise_gsheet_url(url)
    try:
        r = requests.get(fetch_url, timeout=60, allow_redirects=True)
        r.raise_for_status()
    except Exception as e:
        raise RuntimeError(f"Could not download the capacity workbook from {fetch_url}\n{e}")

    data = r.content
    if not data.startswith(_XLSX_MAGIC):
        raise RuntimeError(
            "The capacity link did not return an .xlsx file - it looks like an "
            "HTML page (usually a sign-in or 'no access' page).\n"
            "For a company-restricted Sheet, set up OAuth (CAPACITY_OAUTH_CLIENT) "
            "or a service account (CAPACITY_SA_KEY), or use a local .xlsx path. "
            "See the README.\n"
            f"URL tried: {fetch_url}"
        )
    return io.BytesIO(data)


def _resolve_source(xlsx_path: str | Path | None,
                    sa_key: str | None = None,
                    oauth_client: str | None = None,
                    oauth_token: str | None = None):
    """Return something load_workbook() can open: a local Path, or in-memory
    .xlsx bytes downloaded from a Google Sheet (service account -> OAuth ->
    public, whichever is configured/available)."""
    if xlsx_path is None:
        return Path(__file__).resolve().parent / DEFAULT_FILE
    src = str(xlsx_path)
    if _is_url(src):
        sheet_id = _extract_sheet_id(src)
        if sheet_id:
            key = _find_sa_key(sa_key)
            if key:
                return _export_gsheet_via_service_account(sheet_id, key)
            client = _find_oauth_client(oauth_client)
            if client:
                token = _default_token_path(client, oauth_token)
                return _export_gsheet_via_oauth(sheet_id, client, token)
            # nothing configured -> try the public path (raises if not public)
        return _fetch_workbook_bytes(src)
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(
            f"Capacity workbook not found: {path}\n"
            f"Set CAPACITY_XLSX to a local .xlsx path or a shared workbook URL. "
            f"The workbook needs a 'Settings' sheet and a 'Capacity' sheet."
        )
    return path


def load_dataframe(xlsx_path: str | Path | None = None,
                   ctx: dict[str, Any] | None = None,
                   sa_key: str | None = None,
                   oauth_client: str | None = None,
                   oauth_token: str | None = None) -> pd.DataFrame:
    """Return Member | Activity | Sprint cap from the capacity workbook.

    `xlsx_path` may be a local path or an http(s) URL. `sa_key` / `oauth_client`
    / `oauth_token` are optional credential paths for reading a Google Sheet.
    `ctx` is accepted and ignored so the signature matches the loader swap.
    """
    source = _resolve_source(xlsx_path, sa_key=sa_key,
                             oauth_client=oauth_client, oauth_token=oauth_token)
    src_name = getattr(source, "name", None) or "capacity workbook"

    wb = load_workbook(source, data_only=False)
    if "Settings" not in wb.sheetnames or "Capacity" not in wb.sheetnames:
        raise ValueError(
            f"{src_name} must have a 'Settings' sheet (Working days in B5, Team "
            f"days off in B6) and a 'Capacity' sheet (Team, Member, Activity, "
            f"Capacity/day, Days off, Sprint Capacity)."
        )

    st = wb["Settings"]
    workdays = _num(st["B5"].value)
    team_off = _num(st["B6"].value)

    cap = wb["Capacity"]
    rows: list[dict[str, Any]] = []
    for r in cap.iter_rows(min_row=2, values_only=True):
        cells = (list(r) + [None] * 6)[:6]
        team, member, activity, perday, daysoff, _computed = cells
        if not member or str(member).strip().upper() == "TOTAL":
            continue
        sprint_cap = max(0.0, _num(perday) * (workdays - team_off - _num(daysoff)))
        rows.append({
            "Member":     str(member).strip(),
            "Activity":   (activity or "Development"),
            "Sprint cap": round(sprint_cap, 2),
        })

    return pd.DataFrame(rows, columns=["Member", "Activity", "Sprint cap"])


if __name__ == "__main__":
    df = load_dataframe()
    if df.empty:
        print("No capacity rows found.")
    else:
        print(df.to_string(index=False))
        print("\nPer-member totals:")
        totals = df.groupby("Member")["Sprint cap"].sum().sort_values(ascending=False)
        print(totals.to_string())
        print(f"\nTeam total: {df['Sprint cap'].sum():.1f}h")
